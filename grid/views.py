"""Дашборд с торговыми графиками.

Страницы:
    /dashboard/                 — список стратегий
    /dashboard/<id>/            — графики по стратегии (свечи+сетка, PnL)
    /dashboard/<id>/data.json   — данные для графиков (для авто-обновления)
"""
import time
from datetime import datetime, timezone as dt_timezone
from decimal import Decimal

from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils.dateparse import parse_date

from grid.models import GridStrategy, GridOrder, Side, StrategyStatus
from grid.services import okx_client as okx


def _f(value) -> float:
    return float(value) if value is not None else None


# Кэш свечей: при обновлении дашборда раз в секунду незачем тянуть 200 свечей
# с биржи каждый раз — свечи меняются медленнее таймфрейма. Цена (last) берётся
# «вживую» отдельно, поэтому кэш свечей не влияет на актуальность графика цены.
_CANDLE_CACHE: dict = {}
_CANDLE_TTL = 8.0  # секунд


def _candles(inst_id: str, bar: str = "1H", limit: int = 200) -> list[dict]:
    """Свечи с биржи (от старых к новым), с кэшем TTL. При ошибке — пустой список."""
    key = (inst_id, bar, limit)
    cached = _CANDLE_CACHE.get(key)
    now = time.time()
    if cached and now - cached[0] < _CANDLE_TTL:
        return cached[1]
    try:
        rows = okx.unwrap(okx.market_api().get_candlesticks(inst_id, bar=bar, limit=str(limit)))
    except Exception:  # noqa: BLE001 — график должен рендериться и без биржи
        return cached[1] if cached else []  # при сбое отдаём прошлый кэш
    out = []
    for r in reversed(rows):  # OKX отдаёт новейшие первыми
        ts = datetime.fromtimestamp(int(r[0]) / 1000, tz=dt_timezone.utc)
        out.append({
            "t": ts.isoformat(),
            "o": float(r[1]), "h": float(r[2]), "l": float(r[3]), "c": float(r[4]),
        })
    _CANDLE_CACHE[key] = (now, out)
    return out


def _pnl_series(strategy: GridStrategy) -> list[dict]:
    """Кумулятивный реализованный PnL по сделкам (та же логика, что в движке)."""
    base = Decimal("0")
    avg = Decimal("0")
    pnl = Decimal("0")
    series = []
    for t in strategy.trades.order_by("ts"):
        if t.side == Side.BUY:
            new_base = base + t.fill_size
            if new_base > 0:
                avg = (avg * base + t.fill_price * t.fill_size) / new_base
            base = new_base
        else:
            closing = min(t.fill_size, base) if base > 0 else Decimal("0")
            if closing > 0 and avg > 0:
                pnl += (t.fill_price - avg) * closing
            base -= t.fill_size
            if base <= 0:
                base = Decimal("0")
                avg = Decimal("0")
        series.append({"t": t.ts.isoformat(), "pnl": float(pnl)})
    return series


def build_chart_data(strategy: GridStrategy, bar: str = "1H") -> dict:
    """Полный набор данных для графиков по стратегии."""
    levels = [
        {"index": lv.index, "price": _f(lv.price), "side": lv.side, "status": lv.status}
        for lv in strategy.grid_levels.order_by("index")
    ]
    trades = [
        {"t": t.ts.isoformat(), "side": t.side, "price": _f(t.fill_price), "size": _f(t.fill_size)}
        for t in strategy.trades.order_by("ts")
    ]
    orders = [
        {
            "index": o.level.index if o.level else None,
            "side": o.side,
            "side_display": o.get_side_display(),
            "price": _f(o.price),
            "size": _f(o.size),
            "filled": _f(o.filled_size),
            "state": o.state,
            "state_display": o.get_state_display(),
            "ord_id": o.ord_id,
            "created": o.created_at.isoformat(),
        }
        for o in strategy.orders.select_related("level").order_by("-created_at")
    ]
    try:
        current = float(okx.get_last_price(strategy.inst_id))
    except Exception:  # noqa: BLE001
        current = None

    pos = getattr(strategy, "position", None)
    return {
        "strategy": {
            "id": strategy.id,
            "name": strategy.name,
            "inst_id": strategy.inst_id,
            "status": strategy.get_status_display(),
            "running": strategy.status == StrategyStatus.RUNNING,
            "p_min": _f(strategy.p_min),
            "p_max": _f(strategy.p_max),
            "stop_loss": _f(strategy.effective_stop_loss),
            "current_price": current,
        },
        "candles": _candles(strategy.inst_id, bar=bar),
        "levels": levels,
        "trades": trades,
        "orders": orders,
        "pnl": _pnl_series(strategy),
        "stats": {
            "orders_live": strategy.orders.filter(state="live").count(),
            "orders_filled": strategy.orders.filter(state="filled").count(),
            "trades_buy": strategy.trades.filter(side=Side.BUY).count(),
            "trades_sell": strategy.trades.filter(side=Side.SELL).count(),
            "position_qty": _f(pos.base_qty) if pos else 0,
            "position_avg": _f(pos.avg_price) if pos else 0,
            "realized_pnl": _f(pos.realized_pnl) if pos else 0,
        },
    }


def dashboard_index(request):
    mode = request.GET.get("mode", "")
    all_strategies = list(GridStrategy.objects.values_list('id', 'mode'))
    demo_strats = [s for s in all_strategies if s[1] == "demo"]
    live_strats = [s for s in all_strategies if s[1] == "live"]

    counts = {
        "all": len(all_strategies),
        "demo": len(demo_strats),
        "live": len(live_strats),
    }

    strategies = GridStrategy.objects.all()
    if mode in ("demo", "live"):
        strategies = strategies.filter(mode=mode)

    return render(request, "grid/dashboard.html",
                  {"strategies": strategies, "mode": mode, "counts": counts})


def strategy_chart(request, pk: int):
    strategy = get_object_or_404(GridStrategy, pk=pk)
    bar = request.GET.get("bar", "1H")
    return render(request, "grid/strategy_chart.html", {"strategy": strategy, "bar": bar})


def strategy_chart_data(request, pk: int):
    strategy = get_object_or_404(GridStrategy, pk=pk)
    bar = request.GET.get("bar", "1H")
    return JsonResponse(build_chart_data(strategy, bar=bar))


# Закрытые сделки = завершённые ордера, которые были на бирже (исполнены/отменены).
# Отклонённые биржей (failed) — это ошибки размещения, не сделки, их не показываем.
_CLOSED_STATES = ("filled", "canceled")


def trades_page(request):
    """Отдельная страница с таблицей торгов (фильтры + группировка)."""
    return render(request, "grid/trades.html", {})


def closed_trades_data(request):
    """Все закрытые сделки (исполненные/отменённые ордера) по всем стратегиям."""
    qs = (GridOrder.objects
          .filter(state__in=_CLOSED_STATES)
          .select_related("strategy", "level")
          .order_by("-created_at")[:2000])
    rows = []
    for o in qs:
        filled = o.filled_size or Decimal("0")
        value = (o.avg_px or o.price) * filled if filled else Decimal("0")
        rows.append({
            "ts": o.created_at.isoformat(),
            "pair": o.strategy.inst_id,
            "strategy": o.strategy.name,
            "strategy_type": o.strategy.strategy_type,
            "type_display": o.strategy.get_strategy_type_display(),
            "mode": o.strategy.mode,
            "mode_display": o.strategy.get_mode_display(),
            "side": o.side,
            "side_display": o.get_side_display(),
            "price": _f(o.price),
            "size": _f(o.size),
            "filled": _f(filled),
            "value": _f(value),
            "state": o.state,
            "state_display": o.get_state_display(),
        })
    states = {r["state"]: r["state_display"] for r in rows}
    types = {r["strategy_type"]: r["type_display"] for r in rows}
    return JsonResponse({
        "orders": rows,
        "pairs": sorted({r["pair"] for r in rows}),
        "strategies": sorted({r["strategy"] for r in rows}),
        "states": [{"value": k, "label": v} for k, v in sorted(states.items())],
        "types": [{"value": k, "label": v} for k, v in sorted(types.items())],
    })


def _filtered_closed_orders(params):
    """Закрытые ордера с применением фильтров из GET (для экспорта)."""
    qs = (GridOrder.objects
          .filter(state__in=_CLOSED_STATES)
          .select_related("strategy"))
    if params.get("pair"):
        qs = qs.filter(strategy__inst_id=params["pair"])
    if params.get("strategy"):
        qs = qs.filter(strategy__name=params["strategy"])
    if params.get("type"):
        qs = qs.filter(strategy__strategy_type=params["type"])
    if params.get("mode"):
        qs = qs.filter(strategy__mode=params["mode"])
    if params.get("side"):
        qs = qs.filter(side=params["side"])
    if params.get("state"):
        qs = qs.filter(state=params["state"])
    if params.get("from") and (d := parse_date(params["from"])):
        qs = qs.filter(created_at__date__gte=d)
    if params.get("to") and (d := parse_date(params["to"])):
        qs = qs.filter(created_at__date__lte=d)
    return qs.order_by("-created_at")[:20000]


def export_closed_trades_xlsx(request):
    """Экспорт отфильтрованных сделок в Excel (.xlsx) + итог заработка (нетто)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("Экспорт недоступен: не установлен openpyxl "
                            "(добавьте в requirements и пересоберите).", status=500)

    orders = _filtered_closed_orders(request.GET)

    wb = Workbook()
    ws = wb.active
    ws.title = "Сделки"
    headers = ["Время", "Пара", "Тип стратегии", "Режим", "Стратегия", "Сторона",
               "Цена", "Объём", "Исполнено", "Сумма, USDT", "Статус"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1A5276")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    buy_sum = sell_sum = vol_sum = 0.0
    for o in orders:
        filled = float(o.filled_size or 0)
        value = float((o.avg_px or o.price) or 0) * filled
        ws.append([
            o.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            o.strategy.inst_id,
            o.strategy.get_strategy_type_display(),
            o.strategy.get_mode_display(),
            o.strategy.name,
            o.get_side_display(),
            float(o.price or 0),
            float(o.size or 0),
            filled,
            round(value, 4),
            o.get_state_display(),
        ])
        vol_sum += filled
        if o.side == Side.SELL:
            sell_sum += value
        else:
            buy_sum += value

    earnings = sell_sum - buy_sum
    ws.append([])
    summary = [
        ("Σ покупки (USDT)", round(buy_sum, 4)),
        ("Σ продажи (USDT)", round(sell_sum, 4)),
        ("Σ объём (исполнено)", round(vol_sum, 8)),
        ("Заработок (нетто = продажи − покупки), USDT", round(earnings, 4)),
    ]
    for label, val in summary:
        ws.append(["", "", "", "", "", "", "", "", "", label, val])
        ws.cell(row=ws.max_row, column=10).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=11).font = Font(
            bold=True, color=("1E7E45" if val >= 0 else "C0392B") if "Заработок" in label else "000000")

    # ширины столбцов
    widths = [20, 12, 18, 8, 22, 10, 14, 14, 14, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    fname = "trades_" + datetime.now(dt_timezone.utc).strftime("%Y%m%d_%H%M%S") + ".xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


def test_page(request):
    """Тестовая страница для проверки API и системы."""
    orders = GridOrder.objects.filter(state__in=_CLOSED_STATES).count()
    return HttpResponse(f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Cripto - Тест</title>
        <style>
            body {{ font-family: monospace; background: #0e1117; color: #e6e6e6; padding: 20px; }}
            .ok {{ color: #3fdc7a; }} .err {{ color: #ff6b6b; }}
        </style>
    </head>
    <body>
        <h1>🧪 Тест системы Cripto</h1>
        <p><span class="ok">✅</span> Сервер работает</p>
        <p><span class="ok">✅</span> БД доступна</p>
        <p><span class="ok">✅</span> Ордеров в БД: <b>{orders}</b></p>
        <p><a href="/dashboard/trades/">→ Таблица торгов</a></p>
        <p><a href="/dashboard/closed-trades.json">→ API JSON</a></p>
        <hr>
        <script>
            fetch('/dashboard/closed-trades.json')
                .then(r => r.json())
                .then(d => {{
                    document.body.innerHTML += '<p><span class="ok">✅</span> API работает: ' + d.orders.length + ' ордеров</p>';
                }})
                .catch(e => {{
                    document.body.innerHTML += '<p><span class="err">❌</span> API ошибка: ' + e.message + '</p>';
                }});
        </script>
    </body>
    </html>
    """, content_type="text/html; charset=utf-8")


def trades_page_simple(request):
    """Упрощённая таблица торгов (без фильтров) для быстрой отладки."""
    orders = GridOrder.objects.filter(state__in=_CLOSED_STATES).select_related('strategy').order_by('-created_at')[:500]

    rows = []
    for o in orders:
        ts = o.created_at.strftime('%H:%M:%S')
        side_cls = 'buy' if o.side == 'buy' else 'sell'
        state_cls = 'filled' if o.state == 'filled' else 'canceled'
        filled = o.filled_size or 0
        price = o.price or 0
        size = o.size or 0
        rows.append(f'<tr><td>{ts}</td><td>{o.strategy.name}</td><td>{o.strategy.inst_id}</td>'
                   f'<td class="{side_cls}">{o.get_side_display()}</td>'
                   f'<td style="text-align:right">{price:.2f}</td>'
                   f'<td style="text-align:right">{size:.6f}</td>'
                   f'<td style="text-align:right">{filled:.6f}</td>'
                   f'<td class="{state_cls}">{o.get_state_display()}</td></tr>')

    html = ('<!DOCTYPE html><html><head><meta charset="utf-8"><title>Таблица торгов</title>'
            '<style>body{font-family:monospace;background:#0e1117;color:#e6e6e6;padding:20px}'
            'table{width:100%;border-collapse:collapse}'
            'th,td{padding:8px;border-bottom:1px solid #232a33;text-align:left}'
            'th{background:#1b212b}.buy{color:#3fdc7a}.sell{color:#ff6b6b}</style>'
            '</head><body><h1>📒 Таблица торгов</h1>'
            f'<p>Закрытые сделки: <b>{len(orders)}</b></p>'
            '<table><tr><th>Время</th><th>Стратегия</th><th>Пара</th><th>Сторона</th>'
            '<th>Цена</th><th>Объём</th><th>Исполнено</th><th>Статус</th></tr>'
            + ''.join(rows) +
            '</table><hr><p><a href="/dashboard/">← Дашборд</a></p></body></html>')

    return HttpResponse(html, content_type="text/html; charset=utf-8")
