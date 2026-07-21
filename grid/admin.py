"""Админка для управления сеточными стратегиями.

Действия на странице стратегий выполняют шаги из документа:
    * проверка подключения к OKX
    * синхронизация параметров инструмента (tickSz/lotSz/minSz)
    * расчёт уровней сетки
    * размещение начальной сетки
    * штатная остановка (отмена всех ордеров)
"""
from django import forms
from django.contrib import admin, messages
from django.db.models import Sum
from django.http import HttpResponseRedirect
from django.shortcuts import redirect, render
from django.urls import path, reverse
from django.utils.html import format_html, format_html_join

import json
import os
from decimal import Decimal
from datetime import datetime, timezone as dt_timezone, timedelta

from django.db.models import Q
from django.utils import timezone as dj_timezone

from django.conf import settings
from django.http import FileResponse, Http404

from grid.forms import QuickStrategyForm
from grid.models import (
    GridStrategy, GridLevel, GridOrder, Trade, Position, StrategyLog, Instrument,
    Document, Service, RiskSettings, EquitySnapshot, GridType, StrategyType,
    StrategyStatus, WorkerStatus,
)
from grid.services import okx_client as okx
from grid.services import risk
from grid.services import runner
from grid.services import supervisor
from grid.services import service_api
from grid.services.builder import (
    create_strategy_for_pair, create_typed_strategy, DEFAULT_PARAMS,
)
from grid.services.grid_engine import GridEngine
from grid.services.instruments import refresh_instruments


# --- инлайны -----------------------------------------------------------------
class GridLevelInline(admin.TabularInline):
    model = GridLevel
    extra = 0
    fields = ("index", "price", "side", "status", "active_order")
    readonly_fields = ("index", "price", "side", "status", "active_order")
    ordering = ("index",)
    can_delete = False
    show_change_link = True


class PositionInline(admin.StackedInline):
    model = Position
    extra = 0
    readonly_fields = ("base_qty", "avg_price", "realized_pnl", "updated_at")
    can_delete = False


class StrategyLogInline(admin.TabularInline):
    model = StrategyLog
    extra = 0
    fields = ("created_at", "level", "message")
    readonly_fields = ("created_at", "level", "message")
    ordering = ("-created_at",)
    can_delete = False
    max_num = 0  # только просмотр (добавление через инлайн запрещено)


def _strategies_to_xlsx(queryset):
    """Выгрузка стратегий в Excel (.xlsx) + итоговая сумма заработка. То же, что в торгах."""
    from django.http import HttpResponse
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("Экспорт недоступен: не установлен openpyxl "
                            "(добавьте в requirements и пересоберите).", status=500)
    import io

    qs = queryset.select_related("position")
    wb = Workbook()
    ws = wb.active
    ws.title = "Стратегии"
    headers = ["Название", "Тип", "Режим", "Инструмент", "Статус", "Pmin", "Pmax",
               "Уровней", "Объём ордера", "Заработок, USDT", "Желаемое", "Последний тик"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1A5276")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")

    total = 0.0
    for s in qs:
        try:
            pnl = float(s.position.realized_pnl)
        except Position.DoesNotExist:
            pnl = 0.0
        total += pnl
        ws.append([
            s.name, s.get_strategy_type_display(), s.get_mode_display(), s.inst_id,
            s.get_status_display(),
            float(s.p_min) if s.p_min is not None else None,
            float(s.p_max) if s.p_max is not None else None,
            s.levels,
            float(s.order_size) if s.order_size is not None else None,
            round(pnl, 4),
            s.get_desired_state_display(),
            s.last_tick_at.strftime("%Y-%m-%d %H:%M:%S") if s.last_tick_at else "",
        ])

    ws.append([])
    ws.append(["", "", "", "", "", "", "", "", "Σ Заработок, USDT", round(total, 4)])
    ws.cell(row=ws.max_row, column=9).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=10).font = Font(
        bold=True, color="1E7E45" if total >= 0 else "C0392B")

    widths = [26, 20, 8, 14, 16, 12, 12, 9, 14, 16, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    fname = "strategies_" + datetime.now(dt_timezone.utc).strftime("%Y%m%d_%H%M%S") + ".xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


class GridStrategyAdminForm(forms.ModelForm):
    """Форма для GridStrategyAdmin — валидирует p_max/p_min только для grid типов."""
    class Meta:
        model = GridStrategy
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        strategy_type = cleaned_data.get('strategy_type')
        p_max = cleaned_data.get('p_max')
        p_min = cleaned_data.get('p_min')

        # p_max/p_min требуются только для grid типов
        if strategy_type == StrategyType.GRID:
            if not p_max or not p_min:
                raise forms.ValidationError(
                    "Для типа «Сетка» требуются верхняя и нижняя цены (Pmax, Pmin).")
        return cleaned_data


@admin.register(GridStrategy)
class GridStrategyAdmin(admin.ModelAdmin):
    form = GridStrategyAdminForm
    change_list_template = "admin/grid/gridstrategy/change_list.html"
    list_display = (
        "name", "type_badge", "mode_badge", "inst_id", "status_badge", "runner_badge",
        "earnings_badge", "order_size", "trading_toggle", "open_chart",
    )
    list_filter = ("mode", "strategy_type", "status", "inst_type")
    search_fields = ("name", "inst_id")
    inlines = [PositionInline, GridLevelInline]  # StrategyLogInline скрыт (логи доступны через кнопки)
    readonly_fields = (
        "trading_controls", "params_help", "risk_check", "params_display",
        "tick_sz", "lot_sz", "min_sz", "is_demo",
        "worker_state", "desired_state", "last_tick_at", "last_error",
        "created_at", "updated_at",
    )
    fieldsets = (
        ("Основное", {
            "fields": ("name", "strategy_type", "mode", "inst_id", "inst_type",
                       "td_mode", "leverage", "status", "risk_check", "trading_controls"),
        }),
        ("Параметры стратегии (DCA / Trend / Scalping / Arbitrage)", {
            "fields": ("params", "params_display", "params_help"),
            "description": "JSON-параметры для не-сеточных типов. Для Scalping: кнопка обновит entry_price текущей ценой с OKX.",
        }),
        ("Диапазон сетки (только для типа «Сетка»)", {
            "fields": ("p_max", "p_min", "levels", "grid_type", "order_size"),
            "classes": ("collapse",),
        }),
        ("Параметры инструмента (с биржи)", {
            "fields": ("tick_sz", "lot_sz", "min_sz", "is_demo"),
        }),
        ("Стоп-лосс (сетка)", {"fields": ("stop_loss_enabled", "stop_loss_price"),
                               "classes": ("collapse",)}),
        ("Рабочий цикл (супервизор)", {
            "fields": ("worker_state", "desired_state", "last_tick_at", "last_error"),
            "description": "Желаемое состояние задаётся кнопками; непрерывный цикл ведёт "
                           "воркер run_bots. «Живость» определяется по свежести heartbeat.",
        }),
        ("Служебное", {"fields": ("created_at", "updated_at")}),
    )
    actions = (
        "action_start_trading", "action_stop_trading", "action_stop_all",
        "action_archive", "action_unarchive", "action_export_xlsx",
        "action_reconcile", "action_check_connection",
        "action_sync_instrument", "action_build_levels",
    )

    @admin.action(description="📥 Экспорт в Excel (выбранные)")
    def action_export_xlsx(self, request, queryset):
        return _strategies_to_xlsx(queryset)

    @admin.action(description="🗄 В архив (остановить и скрыть из работы)")
    def action_archive(self, request, queryset):
        n = 0
        for s in queryset.exclude(status=StrategyStatus.ARCHIVED):
            if s.status == StrategyStatus.RUNNING or runner.is_running(s):
                runner.stop_trading(s)  # desired=stop + отмена ордеров
            GridStrategy.objects.filter(pk=s.pk).update(
                status=StrategyStatus.ARCHIVED, desired_state="stop")
            n += 1
        self.message_user(request, f"В архив отправлено стратегий: {n}.", messages.SUCCESS)

    @admin.action(description="♻️ Вернуть из архива")
    def action_unarchive(self, request, queryset):
        n = queryset.filter(status=StrategyStatus.ARCHIVED).update(
            status=StrategyStatus.STOPPED)
        self.message_user(request, f"Возвращено из архива (в «Остановлена»): {n}.",
                          messages.SUCCESS)

    @admin.action(description="🛑 Остановить ВСЕ стратегии (kill-switch)")
    def action_stop_all(self, request, queryset):
        n = risk.stop_all("ручной kill-switch из админки")
        self.message_user(request, f"Kill-switch: остановлено запущенных стратегий — {n}.",
                          messages.WARNING)

    @admin.action(description="🔄 Синхронизировать с биржей (полная сверка)")
    def action_reconcile(self, request, queryset):
        try:
            res = supervisor.reconcile_now()
            self.message_user(
                request,
                f"Сверка завершена: приведены к желаемому состоянию все стратегии; "
                f"отменено осиротевших ордеров — {res['canceled_orphans']}, "
                f"исправлено рассинхронов ордеров — {res['fixed_orders']}.",
                messages.SUCCESS)
        except Exception as e:  # noqa: BLE001
            self.message_user(request, f"Ошибка сверки: {e}", messages.ERROR)

    @admin.display(description="Проверка риска (комиссия / плечо)")
    def risk_check(self, obj):
        if not obj or not obj.pk:
            return "—"
        warnings = risk.fee_step_warnings(obj) + risk.leverage_warnings(obj)
        if not warnings:
            return format_html('<span style="color:#0a0">✓ шаг прибыли выше комиссии, '
                               'плечо/маржа в норме</span>')
        return format_html(
            '<b style="color:#c00">⚠</b> {}',
            format_html_join(format_html("<br>"), "• {}", ((w,) for w in warnings)))

    @admin.display(description="Статус")
    def status_badge(self, obj):
        colors = {
            "draft": "#888", "ready": "#0a7", "running": "#0a0",
            "stopped": "#c80", "emergency": "#c00", "archived": "#6c7a89",
        }
        return format_html(
            '<b style="color:{}">{}</b>', colors.get(obj.status, "#000"),
            obj.get_status_display(),
        )

    @admin.display(description="Цикл")
    def runner_badge(self, obj):
        if runner.is_running(obj):
            return format_html('<b style="color:#0a0" title="свежий heartbeat">● работает</b>')
        if runner.is_stale(obj):
            return format_html(
                '<b style="color:#c00" title="статус RUNNING, но воркер не тикает">'
                '⚠ завис</b>')
        if obj.desired_state == "run":
            return format_html('<span style="color:#c80" title="ждёт воркер run_bots">'
                               '◍ ожидает</span>')
        return format_html('<span style="color:#999">○ остановлен</span>')

    @admin.display(description="Управление")
    def trading_toggle(self, obj):
        """Кнопка запуска/остановки торговли (toggle)."""
        if not obj or not obj.pk:
            return "—"

        if obj.desired_state == "run":
            # Стратегия запущена — показываем кнопку "Остановить"
            return format_html(
                '<a href="javascript:void(0)" onclick="toggleTrading({})" '
                'style="background:#c00; color:white; padding:4px 8px; '
                'border-radius:3px; cursor:pointer; text-decoration:none; '
                'display:inline-block; font-weight:bold">⏹ Стоп</a>',
                obj.pk)
        else:
            # Стратегия остановлена — показываем кнопку "Запустить"
            return format_html(
                '<a href="javascript:void(0)" onclick="toggleTrading({})" '
                'style="background:#0a0; color:white; padding:4px 8px; '
                'border-radius:3px; cursor:pointer; text-decoration:none; '
                'display:inline-block; font-weight:bold">▶ Пуск</a>',
                obj.pk)

    @admin.display(description="Состояние воркера")
    def worker_state(self, obj):
        if not obj or not obj.pk:
            return "—"
        if runner.is_running(obj):
            state = format_html('<b style="color:#0a0">● работает (heartbeat свежий)</b>')
        elif runner.is_stale(obj):
            state = format_html('<b style="color:#c00">⚠ завис — статус RUNNING, но '
                                'воркер run_bots не тикает</b>')
        elif obj.desired_state == "run":
            state = format_html('<span style="color:#c80">◍ ожидает запуска воркером</span>')
        else:
            state = format_html('<span style="color:#999">○ остановлен</span>')
        hb = obj.last_tick_at.strftime("%H:%M:%S") if obj.last_tick_at else "—"
        return format_html('{} &nbsp;·&nbsp; желаемое: <b>{}</b> &nbsp;·&nbsp; '
                           'heartbeat: {}', state, obj.get_desired_state_display(), hb)

    @admin.display(description="Тип")
    def type_badge(self, obj):
        colors = {"grid": "#2471a3", "dca": "#117a3d", "trend": "#8e44ad",
                  "arbitrage": "#b9770e", "scalping": "#a93226"}
        return format_html('<b style="color:{}">{}</b>',
                           colors.get(obj.strategy_type, "#000"),
                           obj.get_strategy_type_display())

    @admin.display(description="Режим")
    def mode_badge(self, obj):
        if obj.mode == "live":
            return format_html('<b style="color:#fff; background:#c0392b; '
                               'padding:1px 7px; border-radius:10px">РЕАЛ</b>')
        return format_html('<span style="color:#fff; background:#7f8c8d; '
                           'padding:1px 7px; border-radius:10px">демо</span>')

    @admin.display(description="Графики")
    def open_chart(self, obj):
        return format_html('<a href="/dashboard/{}/" target="_blank">📈 открыть</a>', obj.id)

    def get_queryset(self, request):
        # select_related по позиции — чтобы колонка «Заработок» не плодила запросы.
        return super().get_queryset(request).select_related("position")

    def changelist_view(self, request, extra_context=None):
        """Добавляет над таблицей суммарный заработок по текущему отбору (фильтры/поиск)."""
        response = super().changelist_view(request, extra_context)
        try:
            cl = response.context_data["cl"]
        except (AttributeError, KeyError, TypeError):
            return response  # редирект/не-табличный ответ
        qs = cl.queryset  # уже с применёнными фильтрами и поиском
        total = qs.aggregate(t=Sum("position__realized_pnl"))["t"] or Decimal("0")
        response.context_data["earnings_total"] = total
        response.context_data["earnings_count"] = qs.count()

        # Признак «воркер не запущен»: есть running-стратегии, но ни у одной нет
        # свежего heartbeat -> run_bots не крутится (иначе он бы обновлял last_tick_at).
        cutoff = dj_timezone.now() - timedelta(seconds=runner.HEARTBEAT_STALE)
        running_qs = GridStrategy.objects.filter(status=StrategyStatus.RUNNING)
        running = running_qs.count()
        fresh = running_qs.filter(last_tick_at__gte=cutoff).count()
        response.context_data["worker_down"] = running > 0 and fresh == 0
        response.context_data["worker_running_cnt"] = running
        return response

    @admin.display(description="Заработок", ordering="position__realized_pnl")
    def earnings_badge(self, obj):
        """Реализованная прибыль стратегии на текущий момент (VWAP), в котируемой валюте."""
        try:
            pnl = obj.position.realized_pnl
        except Position.DoesNotExist:
            pnl = None
        if pnl is None:
            return format_html('<span style="color:#999" title="сделок ещё не было">—</span>')
        color = "#0a0" if pnl > 0 else ("#c00" if pnl < 0 else "#888")
        sign = "+" if pnl > 0 else ""
        return format_html('<b style="color:{}">{}{} USDT</b>', color, sign, f"{pnl:.2f}")

    # Схемы параметров (params) по типам стратегий — подсказка в форме
    PARAM_SCHEMAS = {
        "dca": ('{"mode":"dip", "base_amount":100, "safety_amount":50, '
                '"price_deviation_pct":2, "safety_count":5, "volume_scale":1.5, '
                '"take_profit_pct":3}  ·  для расписания: {"mode":"schedule", '
                '"base_amount":50, "interval_hours":24, "take_profit_pct":3}'),
        "trend": ('{"bar":"1H", "fast":9, "slow":21, "order_amount":100, '
                  '"use_rsi":true, "rsi_period":14, "rsi_overbought":70}'),
        "scalping": '{"order_amount":50, "target_pct":0.3, "stop_pct":0.5}',
        "arbitrage": ('{"base":"USDT", "mid":"BTC", "cross":"ETH", "amount":50, '
                      '"min_profit_pct":0.3, "fee_pct":0.1, "execute":false}'),
        "grid": "Сетка использует отдельные поля ниже (Pmax/Pmin/уровни/объём), params не нужен.",
    }

    @admin.display(description="Схема параметров")
    def params_help(self, obj):
        schema = self.PARAM_SCHEMAS.get(getattr(obj, "strategy_type", "grid"), "")
        return format_html(
            '<div style="font-size:12px; color:#555">Пример params для типа '
            '<b>{}</b>:<br><code style="display:block; background:#f4f5f7; '
            'padding:8px; border-radius:6px; margin-top:4px; white-space:pre-wrap">{}</code>'
            '</div>',
            obj.get_strategy_type_display() if obj and obj.pk else "—", schema,
        )

    @admin.display(description="Управление торговлей")
    def trading_controls(self, obj):
        """Кнопки запуска/остановки торговли и ссылка на графики.

        Кнопки — submit'ы той же формы; их обрабатывает response_change().
        """
        if not obj or not obj.pk:
            return "Сохраните стратегию, чтобы управлять торговлей."
        confirm = ("ВНИМАНИЕ! Реальная торговля на НАСТОЯЩИЕ деньги. "
                   "Стратегия будет запущена в режиме РЕАЛ. Продолжить?")
        btn = "color:#fff; padding:6px 12px; border:0; border-radius:6px; cursor:pointer;"
        logs_url = reverse("admin:grid_strategylog_changelist") + f"?strategy__id__exact={obj.pk}"
        errs_url = logs_url + "&level__exact=error"
        return format_html(
            '<div style="display:flex; gap:10px; align-items:center; flex-wrap:wrap;">'
            '<input type="submit" name="_start_trading" value="▶️ Запустить торговлю" '
            'style="background:#1f7a3d; {btn}">'
            '<input type="submit" name="_stop_trading" value="⏹ Остановить торговлю" '
            'style="background:#a33; {btn}">'
            '<input type="submit" name="_copy_strategy" value="📋 Скопировать стратегию" '
            'style="background:#566573; {btn}">'
            '<input type="submit" name="_start_live" value="🔴 Торговать в реале" '
            'onclick="return confirm(\'{confirm}\')" style="background:#c0392b; {btn}">'
            '<input type="submit" name="_reconcile" value="🔄 Синхронизировать" '
            'style="background:#7d3c98; {btn}">'
            '<a class="button" href="/dashboard/{pk}/" target="_blank" '
            'style="background:#264b7a; color:#fff;">📈 Открыть графики</a>'
            '<a class="button" href="{logs_url}" '
            'style="background:#5d6d7e; color:#fff;">🧾 Логи стратегии</a>'
            '<a class="button" href="{errs_url}" '
            'style="background:#8a5a00; color:#fff;">⚠ Ошибки</a>'
            '</div>',
            btn=btn, confirm=confirm, pk=obj.pk, logs_url=logs_url, errs_url=errs_url,
        )

    # --- кнопки запуска/остановки на странице объекта ------------------------
    def response_change(self, request, obj):
        if "_start_trading" in request.POST:
            self._do(request, obj, runner.start_trading)
            return HttpResponseRedirect(request.path)
        if "_stop_trading" in request.POST:
            self._do(request, obj, runner.stop_trading)
            return HttpResponseRedirect(request.path)
        if "_start_live" in request.POST:
            # переключаем стратегию в РЕАЛ и запускаем
            obj.mode = "live"
            obj.save(update_fields=["mode", "updated_at"])
            self.message_user(request, "⚠ Режим переключён на РЕАЛ.", messages.WARNING)
            self._do(request, obj, runner.start_trading)
            return HttpResponseRedirect(request.path)
        if "_reconcile" in request.POST:
            try:
                supervisor.tick_strategy(obj)
                obj.refresh_from_db()
                self.message_user(
                    request,
                    f"Синхронизировано. Статус: {obj.get_status_display()}, "
                    f"желаемое: {obj.get_desired_state_display()}."
                    + (f" Ошибка: {obj.last_error}" if obj.last_error else ""),
                    messages.SUCCESS if not obj.last_error else messages.WARNING)
            except Exception as e:  # noqa: BLE001
                self.message_user(request, f"Ошибка синхронизации: {e}", messages.ERROR)
            return HttpResponseRedirect(request.path)
        if "_copy_strategy" in request.POST:
            new = self._copy(obj)
            self.message_user(
                request, f"Создана копия «{new.name}» (не запущена).", messages.SUCCESS)
            return HttpResponseRedirect(
                reverse("admin:grid_gridstrategy_change", args=[new.pk]))
        return super().response_change(request, obj)

    def _do(self, request, obj, fn):
        try:
            res = fn(obj)
            self.message_user(
                request, res["msg"],
                messages.SUCCESS if res.get("ok") else messages.WARNING,
            )
        except Exception as e:  # noqa: BLE001
            self.message_user(request, f"Ошибка: {e}", messages.ERROR)

    def _copy(self, obj):
        """Клонирует конфиг стратегии (без запуска). Возвращает новую стратегию."""
        base = f"{obj.name} (копия)"
        name, i = base, 2
        while GridStrategy.objects.filter(name=name).exists():
            name = f"{base} {i}"
            i += 1
        params = dict(obj.params or {})
        params.pop("_state", None)  # рантайм-состояние не копируем
        status = "ready" if obj.tick_sz is not None else "draft"
        return GridStrategy.objects.create(
            name=name, strategy_type=obj.strategy_type, mode=obj.mode,
            inst_id=obj.inst_id, inst_type=obj.inst_type, td_mode=obj.td_mode,
            params=params, p_max=obj.p_max, p_min=obj.p_min, levels=obj.levels,
            grid_type=obj.grid_type, order_size=obj.order_size, tick_sz=obj.tick_sz,
            lot_sz=obj.lot_sz, min_sz=obj.min_sz, stop_loss_enabled=obj.stop_loss_enabled,
            stop_loss_price=obj.stop_loss_price, is_demo=obj.is_demo, status=status,
        )

    # --- быстрое создание стратегии по паре ----------------------------------
    def get_urls(self):
        from django.urls import path as django_path
        custom = [
            path("quick-create/", self.admin_site.admin_view(self.quick_create_view),
                 name="grid_gridstrategy_quick_create"),
            path("export-xlsx/", self.admin_site.admin_view(self.export_xlsx_view),
                 name="grid_gridstrategy_export_xlsx"),
            django_path(
                "<int:pk>/trading-toggle/",
                self.admin_site.admin_view(self.trading_toggle_view),
                name="grid_gridstrategy_trading_toggle",
            ),
            django_path(
                "<int:pk>/update-entry-price/",
                self.admin_site.admin_view(self.update_entry_price_view),
                name="grid_gridstrategy_update_entry_price",
            ),
            django_path(
                "<int:pk>/save-scalping-params/",
                self.admin_site.admin_view(self.save_scalping_params_view),
                name="grid_gridstrategy_save_scalping_params",
            ),
        ]
        return custom + super().get_urls()

    def export_xlsx_view(self, request):
        """Экспорт в Excel текущего отфильтрованного вида списка стратегий."""
        try:
            cl = self.get_changelist_instance(request)
            qs = cl.get_queryset(request)
        except Exception:  # noqa: BLE001 — при кривых параметрах отдаём всё
            qs = self.get_queryset(request)
        return _strategies_to_xlsx(qs)

    def quick_create_view(self, request):
        """Мастер: выбираешь тип и пары -> стратегии создаются с авто-заполнением."""
        if request.method == "POST":
            form = QuickStrategyForm(request.POST)
            if form.is_valid():
                cd = form.cleaned_data
                stype = cd["strategy_type"]
                raw = cd["pairs"].replace(",", "\n").splitlines()
                pairs = [p.strip() for p in raw if p.strip()]
                created = 0
                for pair in pairs:
                    try:
                        if stype == StrategyType.GRID:
                            res = create_strategy_for_pair(
                                pair, range_pct=cd.get("range_pct") or 10,
                                levels=cd.get("levels") or 10,
                                order_notional=cd.get("order_notional") or 15,
                                grid_type=cd.get("grid_type") or GridType.ARITHMETIC,
                            )
                        else:
                            res = create_typed_strategy(pair, stype, cd.get("params") or {})
                    except Exception as e:  # noqa: BLE001
                        self.message_user(request, f"{pair}: ошибка — {e}", messages.ERROR)
                        continue
                    self.message_user(
                        request, res["msg"],
                        messages.SUCCESS if res["ok"] else messages.WARNING,
                    )
                    if res["ok"]:
                        created += 1
                        if cd.get("start"):
                            self._do(request, res["strategy"], runner.start_trading)
                if created:
                    return redirect("admin:grid_gridstrategy_changelist")
        else:
            form = QuickStrategyForm()

        instruments = list(
            Instrument.objects.filter(active=True)
            .order_by("inst_id").values_list("inst_id", flat=True)
        )
        ctx = {
            **self.admin_site.each_context(request),
            "title": "Мастер создания стратегии",
            "opts": self.model._meta,
            "form": form,
            # дефолтные params по типам — для авто-заполнения в форме (JS)
            "defaults_json": json.dumps(
                {str(k): v for k, v in DEFAULT_PARAMS.items()}, ensure_ascii=False),
            # справочник пар для выбора (кнопка «Обновить пары с биржи» в разделе Инструменты)
            "instruments": instruments,
        }
        return render(request, "admin/grid/quick_create.html", ctx)

    # --- массовые действия запуска/остановки --------------------------------
    @admin.action(description="▶️ Запустить торговлю")
    def action_start_trading(self, request, queryset):
        for s in queryset:
            self._do(request, s, runner.start_trading)

    @admin.action(description="⏹ Остановить торговлю")
    def action_stop_trading(self, request, queryset):
        for s in queryset:
            self._do(request, s, runner.stop_trading)

    # --- вспомогательные действия --------------------------------------------
    @admin.action(description="🔌 Проверить подключение к OKX")
    def action_check_connection(self, request, queryset):
        try:
            info = okx.check_connection()
            self.message_user(
                request,
                f"OKX на связи. Демо: {info['demo']}, время сервера: {info['server_time_ms']}.",
                messages.SUCCESS,
            )
        except Exception as e:  # noqa: BLE001
            self.message_user(request, f"OKX недоступен: {e}", messages.ERROR)

    @admin.action(description="📐 Синхронизировать параметры инструмента")
    def action_sync_instrument(self, request, queryset):
        for s in queryset:
            try:
                GridEngine(s).sync_instrument()
                self.message_user(
                    request,
                    f"[{s.name}] tickSz={s.tick_sz}, lotSz={s.lot_sz}, minSz={s.min_sz}",
                    messages.SUCCESS,
                )
            except Exception as e:  # noqa: BLE001
                self.message_user(request, f"[{s.name}] ошибка: {e}", messages.ERROR)

    @admin.action(description="🧮 Рассчитать уровни сетки (только тип «Сетка»)")
    def action_build_levels(self, request, queryset):
        for s in queryset:
            if s.strategy_type != "grid":
                self.message_user(
                    request, f"[{s.name}] расчёт уровней только для сеток — пропущено.",
                    messages.WARNING)
                continue
            try:
                n = GridEngine(s).build_levels()
                self.message_user(request, f"[{s.name}] рассчитано уровней: {n}", messages.SUCCESS)
            except Exception as e:  # noqa: BLE001
                self.message_user(request, f"[{s.name}] ошибка: {e}", messages.ERROR)

    @admin.display(description="Параметры")
    def params_display(self, obj):
        """Красивый вывод параметров JSON с возможностью редактирования для Scalping."""
        if not obj or not obj.params:
            return "—"

        params = obj.params or {}

        # Для Scalping стратегии выводим интерактивный блок
        if obj.strategy_type == 'scalping':
            stop_pct = params.get('stop_pct', 0.5)
            target_pct = params.get('target_pct', 0.3)
            order_amount = params.get('order_amount', 50)
            entry_price = params.get('_state', {}).get('entry_price', 0)

            html = f'''
            <div id="scalping-params-block-{obj.pk}" class="scalping-params-block">
                <!-- Display mode (default) -->
                <div id="scalping-params-display-{obj.pk}">
                    <div class="params-display-row">
                        <label>stop_pct:</label>
                        <div class="value">{stop_pct}</div>
                    </div>
                    <div class="params-display-row">
                        <label>target_pct:</label>
                        <div class="value">{target_pct}</div>
                    </div>
                    <div class="params-display-row">
                        <label>order_amount:</label>
                        <div class="value">{order_amount}</div>
                    </div>
                    <div class="params-display-row">
                        <label>entry_price:</label>
                        <div class="value">{entry_price}</div>
                    </div>
                    <div class="params-buttons">
                        <button type="button" class="params-button primary" onclick="toggleEditMode({obj.pk})">✏️ Редактировать</button>
                        <button type="button" class="params-button secondary" onclick="updateEntryPrice({obj.pk})">🔄 Цена OKX</button>
                    </div>
                </div>

                <!-- Edit mode (hidden by default) -->
                <div id="scalping-params-edit-{obj.pk}" style="display:none;">
                    <div class="params-row">
                        <label for="stop_pct-{obj.pk}">stop_pct:</label>
                        <input type="number" id="stop_pct-{obj.pk}" value="{stop_pct}" step="0.01" min="0" max="100">
                    </div>
                    <div class="params-row">
                        <label for="target_pct-{obj.pk}">target_pct:</label>
                        <input type="number" id="target_pct-{obj.pk}" value="{target_pct}" step="0.01" min="0" max="100">
                    </div>
                    <div class="params-row">
                        <label for="order_amount-{obj.pk}">order_amount:</label>
                        <input type="number" id="order_amount-{obj.pk}" value="{order_amount}" step="0.01" min="0">
                    </div>
                    <div class="params-row">
                        <label for="entry_price-{obj.pk}">entry_price:</label>
                        <input type="number" id="entry_price-{obj.pk}" value="{entry_price}" step="0.01" min="0">
                    </div>
                    <div class="params-buttons">
                        <button type="button" class="params-button primary" onclick="saveScalpingParams({obj.pk})">💾 Сохранить</button>
                        <button type="button" class="params-button secondary" onclick="toggleEditMode({obj.pk})">✕ Отмена</button>
                    </div>
                </div>
            </div>
            '''
            return format_html(html)

        # Для остальных типов просто выводим текст
        lines = []
        for key in ['stop_pct', 'target_pct', 'order_amount']:
            val = params.get(key)
            if val is not None:
                lines.append(f"<b>{key}:</b> {val}")

        entry = params.get('_state', {}).get('entry_price')
        if entry:
            lines.append(f"<b>entry_price:</b> {entry}")

        return format_html('<br>'.join(lines))

    def save_scalping_params_view(self, request, pk):
        """Сохраняет параметры Scalping."""
        from django.http import JsonResponse
        import json
        try:
            s = GridStrategy.objects.get(pk=pk, strategy_type='scalping')
            data = json.loads(request.body)

            # Обновляем params
            params = s.params or {}
            params['stop_pct'] = float(data.get('stop_pct', params.get('stop_pct', 0.5)))
            params['target_pct'] = float(data.get('target_pct', params.get('target_pct', 0.3)))
            params['order_amount'] = float(data.get('order_amount', params.get('order_amount', 50)))
            if '_state' not in params:
                params['_state'] = {}
            params['_state']['entry_price'] = float(data.get('entry_price', params.get('_state', {}).get('entry_price', 0)))

            s.params = params
            s.save(update_fields=["params"])

            return JsonResponse({
                "ok": True,
                "params": params
            })
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)}, status=400)

    def update_entry_price_view(self, request, pk):
        """Обновляет entry_price из текущей цены OKX."""
        from django.http import JsonResponse
        try:
            s = GridStrategy.objects.get(pk=pk, strategy_type='scalping')
            current_price = okx.get_last_price(s.inst_id)

            # Обновляем params
            params = s.params or {}
            if '_state' not in params:
                params['_state'] = {}
            params['_state']['entry_price'] = float(current_price)
            s.params = params
            s.save(update_fields=["params"])

            return JsonResponse({
                "ok": True,
                "entry_price": float(current_price),
                "inst_id": s.inst_id
            })
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)}, status=400)

    def trading_toggle_view(self, request, pk):
        """Переключает desired_state стратегии (run ↔ stop)."""
        from django.http import JsonResponse
        try:
            s = GridStrategy.objects.get(pk=pk)
            s.desired_state = "stop" if s.desired_state == "run" else "run"
            s.save(update_fields=["desired_state"])
            return JsonResponse({"ok": True, "desired_state": s.desired_state})
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)}, status=400)


@admin.register(GridLevel)
class GridLevelAdmin(admin.ModelAdmin):
    list_display = ("strategy", "index", "price", "side", "status", "active_order")
    list_filter = ("status", "side", "strategy")
    search_fields = ("strategy__name",)
    ordering = ("strategy", "index")


@admin.register(GridOrder)
class GridOrderAdmin(admin.ModelAdmin):
    list_display = (
        "strategy", "side", "price", "size", "state",
        "filled_size", "ord_id", "created_at",
    )
    list_filter = ("state", "side", "strategy")
    search_fields = ("ord_id", "cl_ord_id", "strategy__name")
    readonly_fields = ("raw", "created_at", "updated_at")
    date_hierarchy = "created_at"


@admin.register(Trade)
class TradeAdmin(admin.ModelAdmin):
    list_display = ("strategy", "side", "fill_price", "fill_size", "fee", "ts")
    list_filter = ("side", "strategy")
    search_fields = ("trade_id", "strategy__name")
    date_hierarchy = "ts"


@admin.register(Position)
class PositionAdmin(admin.ModelAdmin):
    list_display = ("strategy", "base_qty", "avg_price", "realized_pnl", "updated_at")
    search_fields = ("strategy__name",)


@admin.register(StrategyLog)
class StrategyLogAdmin(admin.ModelAdmin):
    list_display = ("created_at", "strategy", "level", "message")
    list_filter = ("level", "strategy")
    search_fields = ("message", "strategy__name")
    date_hierarchy = "created_at"


@admin.register(Instrument)
class InstrumentAdmin(admin.ModelAdmin):
    """Справочник торговых пар с биржи + кнопка обновления."""
    change_list_template = "admin/grid/instrument/change_list.html"
    list_display = ("inst_id", "active", "base_ccy", "quote_ccy", "min_sz",
                    "tick_sz", "state", "updated_at")
    list_editable = ("active",)
    list_filter = ("active", "inst_type", "quote_ccy", "state")
    search_fields = ("inst_id", "base_ccy", "quote_ccy")
    ordering = ("inst_id",)
    list_per_page = 50
    actions = ("make_active", "make_inactive")

    @admin.action(description="✅ Сделать Active")
    def make_active(self, request, queryset):
        n = queryset.update(active=True)
        self.message_user(request, f"Активировано пар: {n}", messages.SUCCESS)

    @admin.action(description="🚫 Снять Active")
    def make_inactive(self, request, queryset):
        n = queryset.update(active=False)
        self.message_user(request, f"Деактивировано пар: {n}", messages.WARNING)

    def get_urls(self):
        custom = [
            path("refresh/", self.admin_site.admin_view(self.refresh_view),
                 name="grid_instrument_refresh"),
        ]
        return custom + super().get_urls()

    def refresh_view(self, request):
        """Кнопка «Обновить пары с биржи»: тянет инструменты с OKX."""
        try:
            res = refresh_instruments("SPOT")
            self.message_user(request, res["msg"], messages.SUCCESS)
        except Exception as e:  # noqa: BLE001
            self.message_user(request, f"Ошибка обновления: {e}", messages.ERROR)
        return redirect("admin:grid_instrument_changelist")


DOCS_DIR = settings.BASE_DIR / "docs"


@admin.register(Document)
class DocumentAdmin(admin.ModelAdmin):
    """Пункт «Документы»: динамический список PDF из папки docs/."""

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return True

    def get_model_perms(self, request):
        # Пустые права -> модель НЕ показывается в группе Grid (это отдельный
        # пункт меню). URL страницы при этом остаётся рабочим.
        return {}

    def get_urls(self):
        return [
            path("", self.admin_site.admin_view(self.list_view),
                 name="grid_document_changelist"),
            path("open/<str:filename>/", self.admin_site.admin_view(self.serve_view),
                 name="grid_document_open"),
        ]

    def list_view(self, request):
        files = []
        if DOCS_DIR.exists():
            for f in sorted(DOCS_DIR.glob("*.pdf")):
                stt = f.stat()
                files.append({
                    "name": f.name,
                    "size_kb": round(stt.st_size / 1024),
                    "mtime": datetime.fromtimestamp(stt.st_mtime, tz=dt_timezone.utc),
                })
        ctx = {
            **self.admin_site.each_context(request),
            "title": "Документы",
            "opts": self.model._meta,
            "files": files,
            "docs_dir": str(DOCS_DIR),
        }
        return render(request, "admin/grid/documents.html", ctx)

    def serve_view(self, request, filename):
        # безопасность: только имя файла (без путей), только .pdf, только из docs/
        name = os.path.basename(filename)
        path_ = (DOCS_DIR / name).resolve()
        if (not name.lower().endswith(".pdf") or not path_.exists()
                or DOCS_DIR.resolve() not in path_.parents):
            raise Http404("Документ не найден")
        as_attach = request.GET.get("download") == "1"
        return FileResponse(open(path_, "rb"), content_type="application/pdf",
                            as_attachment=as_attach, filename=name)


@admin.register(Service)
class ServiceAdmin(admin.ModelAdmin):
    """Раздел «Сервис (API)»: страницы «Демо» и «Реал» с read-only запросами к OKX."""

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return True

    def get_model_perms(self, request):
        # Пустые права -> модель не показывается в группе Grid (это отдельный
        # раздел меню с двумя пунктами Демо/Реал). URL-адреса при этом рабочие.
        return {}

    def get_urls(self):
        return [
            path("", self.admin_site.admin_view(self.redirect_view),
                 name="grid_service_changelist"),
            path("demo/", self.admin_site.admin_view(self.demo_view),
                 name="grid_service_demo"),
            path("real/", self.admin_site.admin_view(self.real_view),
                 name="grid_service_real"),
        ]

    def redirect_view(self, request):
        return redirect("admin:grid_service_demo")

    def demo_view(self, request):
        return self._console(request, "demo")

    def real_view(self, request):
        return self._console(request, "live")

    def _console(self, request, mode):
        inst = request.GET.get("inst", service_api.DEFAULT_INST)
        key = request.GET.get("q", "")
        result = service_api.run(mode, key, inst) if key else None
        is_live = mode == "live"
        ctx = {
            **self.admin_site.each_context(request),
            "title": f"Сервис (API) — {'РЕАЛ' if is_live else 'Демо'}",
            "opts": self.model._meta,
            "mode": mode,
            "is_live": is_live,
            "catalog": service_api.CATALOG,
            "inst": inst,
            "active_key": key,
            "result": result,
            "demo_url": reverse("admin:grid_service_demo"),
            "real_url": reverse("admin:grid_service_real"),
        }
        return render(request, "admin/grid/service_console.html", ctx)


@admin.register(RiskSettings)
class RiskSettingsAdmin(admin.ModelAdmin):
    """Глобальные риск-настройки — синглтон (одна запись)."""

    def has_add_permission(self, request):
        return not RiskSettings.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        # синглтон: сразу открываем единственную запись на редактирование
        obj = RiskSettings.load()
        return HttpResponseRedirect(
            reverse("admin:grid_risksettings_change", args=[obj.pk]))


@admin.register(EquitySnapshot)
class EquitySnapshotAdmin(admin.ModelAdmin):
    list_display = ("ts", "equity")
    date_hierarchy = "ts"
    ordering = ("-ts",)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False


# =========================================================================
#  WorkerStatus — мониторинг воркера в специальной группе админки
# =========================================================================
@admin.register(WorkerStatus)
class WorkerStatusAdmin(admin.ModelAdmin):
    """Раздел мониторинга статуса воркера с информацией о его работе."""

    list_display = (
        "worker_status_display", "last_heartbeat", "strategies_info",
        "cycles_info", "error_display"
    )
    list_filter = ("is_running",)
    readonly_fields = (
        "worker_status_display", "last_heartbeat", "strategies_count",
        "running_count", "stopped_count", "orders_processed", "cycles_completed",
        "error_display", "is_running"
    )
    fieldsets = (
        ("🟢 Статус воркера", {
            "fields": ("worker_status_display", "is_running", "last_heartbeat")
        }),
        ("📊 Статистика стратегий", {
            "fields": ("strategies_count", "running_count", "stopped_count")
        }),
        ("⚙️ Статистика обработки", {
            "fields": ("orders_processed", "cycles_completed")
        }),
        ("🚨 Ошибки", {
            "fields": ("error_display",),
            "classes": ("collapse",)
        }),
    )

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.filter(pk=1)

    @admin.display(description="Статус")
    def worker_status_display(self, obj):
        """Отображение статуса воркера с иконкой."""
        if obj.is_running:
            return format_html(
                '<span style="color:#0a0; font-weight:bold; font-size:14px;">🟢 Воркер работает</span>'
            )
        else:
            return format_html(
                '<span style="color:#c00; font-weight:bold; font-size:14px;">🔴 Воркер остановлен</span>'
            )

    @admin.display(description="Последний heartbeat")
    def last_heartbeat(self, obj):
        """Время последнего heartbeat с расчетом прошедшего времени."""
        from datetime import datetime, timezone as dt_timezone, timedelta

        elapsed = (datetime.now(dt_timezone.utc) - obj.last_heartbeat).total_seconds()
        elapsed_str = f"{elapsed:.0f}"

        if elapsed < 30:
            return format_html(
                '<span style="color:#0a0; font-weight:bold;">✓ {}с назад</span>',
                elapsed_str
            )
        elif elapsed < 120:
            return format_html(
                '<span style="color:#fa0; font-weight:bold;">⚠ {}с назад</span>',
                elapsed_str
            )
        else:
            return format_html(
                '<span style="color:#c00; font-weight:bold;">✗ {}с назад</span>',
                elapsed_str
            )

    @admin.display(description="Стратегии")
    def strategies_info(self, obj):
        """Информация о количестве стратегий."""
        return format_html(
            '<span style="font-weight:bold;">Всего: {}</span><br/>'
            '<span style="color:#0a0;">▶ Запущено: {}</span><br/>'
            '<span style="color:#666;">⏸ Остановлено: {}</span>',
            obj.strategies_count, obj.running_count, obj.stopped_count
        )

    @admin.display(description="Обработано")
    def cycles_info(self, obj):
        """Информация о циклах и ордерах."""
        cycles_str = f"{obj.cycles_completed:,}".replace(',', ' ')
        orders_str = f"{obj.orders_processed:,}".replace(',', ' ')
        return format_html(
            'Циклов: <b>{}</b><br/>Ордеров: <b>{}</b>',
            cycles_str, orders_str
        )

    @admin.display(description="Последняя ошибка")
    def error_display(self, obj):
        """Отображение последней ошибки."""
        if obj.last_error:
            return format_html(
                '<div style="background:#fdd; padding:8px; border-radius:4px; '
                'border:1px solid #faa; font-family:monospace; font-size:12px; '
                'max-width:500px; overflow-x:auto;">{}</div>',
                obj.last_error[:500]
            )
        return format_html('<span style="color:#0a0;">✓ Ошибок не обнаружено</span>')


# Заголовок админки отражает режим: демо или РЕАЛЬНАЯ торговля (боевые ключи).
if okx.is_live():
    admin.site.site_header = "Cripto — ⚠ РЕАЛЬНАЯ ТОРГОВЛЯ (LIVE)"
    admin.site.index_title = "⚠ РЕАЛЬНАЯ ТОРГОВЛЯ — используются боевые ключи (flag=0)"
else:
    admin.site.site_header = f"Cripto — Grid Trading (демо · режим {okx.mode()})"
    admin.site.index_title = "Управление стратегиями · демо-режим"
admin.site.site_title = "Cripto Admin"
# Главная страница админки с дополнительным блоком «Графики торговли»
admin.site.index_template = "admin/custom_index.html"
